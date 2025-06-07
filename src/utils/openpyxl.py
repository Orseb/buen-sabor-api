from datetime import datetime
from io import BytesIO
from typing import List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font


def generate_excel_report(
    buffer: BytesIO,
    purchases: List[Tuple[datetime, int, float, str]],
    invoices: List[Tuple[datetime, float, str]],
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> BytesIO:
    """Generate an Excel report for movements, costs, and revenues."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    headers = [
        "Tipo",
        "Fecha",
        "Descripción",
        "Cantidad",
        "Costo Total",
        "Ganancia Total",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    row = 2
    for purchase_date, quantity, total_cost, name in purchases:
        sheet.cell(row=row, column=1).value = "Compra de Inventario"
        sheet.cell(row=row, column=2).value = purchase_date
        sheet.cell(row=row, column=3).value = name
        sheet.cell(row=row, column=4).value = quantity
        sheet.cell(row=row, column=5).value = total_cost
        sheet.cell(row=row, column=6).value = None
        row += 1

    for invoice_date, total, invoice_type in invoices:
        sheet.cell(row=row, column=1).value = "Venta"
        sheet.cell(row=row, column=2).value = invoice_date
        sheet.cell(row=row, column=3).value = invoice_type.value
        sheet.cell(row=row, column=4).value = None
        sheet.cell(row=row, column=5).value = None
        sheet.cell(row=row, column=6).value = total
        row += 1

    workbook.save(buffer)
